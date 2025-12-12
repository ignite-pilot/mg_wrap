from flask import Blueprint, request, jsonify
from app import db
from app.models import Asset, StorageApplication, AssetCategory, AssetStatus
from app.routes.auth import get_current_user
from app.utils import handle_error
from datetime import datetime
from openpyxl import load_workbook

assets_bp = Blueprint('assets', __name__)

def generate_asset_number():
    """자산번호 생성 (예: ASSET-2024-001)"""
    year = datetime.now().year
    # 같은 년도의 마지막 자산번호 찾기
    last_asset = Asset.query.filter(
        Asset.asset_number.like(f'ASSET-{year}-%')
    ).order_by(Asset.asset_number.desc()).first()
    
    if last_asset:
        try:
            last_num = int(last_asset.asset_number.split('-')[-1])
            new_num = last_num + 1
        except (ValueError, IndexError):
            new_num = 1
    else:
        new_num = 1
    
    return f'ASSET-{year}-{str(new_num).zfill(3)}'

@assets_bp.route('/create', methods=['POST'])
def create_asset():
    """보관 자산 개별 등록"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        data = request.get_json()
        storage_application_id = data.get('storage_application_id')
        application_date = data.get('application_date')
        storage_start_date = data.get('storage_start_date')
        asset_category = data.get('asset_category')
        special_notes = data.get('special_notes', '')
        
        if not all([storage_application_id, application_date, asset_category]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # 보관 신청 확인
        application = StorageApplication.query.get_or_404(storage_application_id)
        if application.user_id != user.get('id'):
            return jsonify({'error': 'Unauthorized'}), 403
        
        # 자산번호 생성
        asset_number = generate_asset_number()
        
        # 카테고리 검증
        try:
            category = AssetCategory(asset_category)
        except ValueError:
            return jsonify({'error': 'Invalid asset category'}), 400
        
        # 날짜 파싱
        app_date = datetime.strptime(application_date, '%Y-%m-%d').date()
        # 보관 시작일이 없으면 None으로 설정
        if storage_start_date:
            start_date = datetime.strptime(storage_start_date, '%Y-%m-%d').date()
        else:
            start_date = None
        
        asset = Asset(
            asset_number=asset_number,
            storage_application_id=storage_application_id,
            application_date=app_date,
            storage_start_date=start_date,
            asset_category=category,
            special_notes=special_notes,
            status=AssetStatus.STORED
        )
        
        db.session.add(asset)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'asset': asset.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@assets_bp.route('/upload', methods=['POST'])
def upload_excel():
    """엑셀 파일로 자산 일괄 등록"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        storage_application_id = request.form.get('storage_application_id')
        
        if not storage_application_id:
            return jsonify({'error': 'storage_application_id is required'}), 400
        
        # 보관 신청 확인
        application = StorageApplication.query.get_or_404(storage_application_id)
        if application.user_id != user.get('id'):
            return jsonify({'error': 'Unauthorized'}), 403
        
        # 엑셀 파일 읽기
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            # 헤더 확인
            headers = [cell.value for cell in ws[1]]
            required_columns = ['보관 신청일', '자산 분류', '특이사항']
            optional_columns = ['보관 시작일']
            header_map = {}
            
            for col in required_columns:
                if col not in headers:
                    return jsonify({'error': f'Missing column: {col}'}), 400
                header_map[col] = headers.index(col) + 1  # 1-based index
            
            for col in optional_columns:
                if col in headers:
                    header_map[col] = headers.index(col) + 1  # 1-based index
            
            assets_created = []
            errors = []
            
            # 데이터 행 처리 (2행부터)
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # 자산번호 생성
                    asset_number = generate_asset_number()
                    
                    # 카테고리 매핑 (한글 -> 영어)
                    category_map = {
                        '사무용품': AssetCategory.OFFICE_SUPPLIES,
                        '서류': AssetCategory.DOCUMENTS,
                        '장비': AssetCategory.EQUIPMENT,
                        '가구': AssetCategory.FURNITURE,
                        '의류': AssetCategory.CLOTHING,
                        '가전': AssetCategory.APPLIANCES,
                        '기타': AssetCategory.OTHER
                    }
                    
                    category_cell = row[header_map['자산 분류'] - 1]
                    category_str = str(category_cell.value).strip() if category_cell.value else ''
                    
                    if not category_str or category_str not in category_map:
                        errors.append(f'Row {idx}: Invalid category "{category_str}"')
                        continue
                    
                    category = category_map[category_str]
                    
                    # 날짜 파싱
                    try:
                        app_date_cell = row[header_map['보관 신청일'] - 1]
                        
                        if app_date_cell.value is None:
                            errors.append(f'Row {idx}: 보관 신청일이 비어있습니다.')
                            continue
                        
                        # 날짜 파싱 (여러 형식 지원)
                        if isinstance(app_date_cell.value, str):
                            # 문자열인 경우 여러 형식 시도
                            date_str = app_date_cell.value.strip()
                            date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']
                            app_date = None
                            for fmt in date_formats:
                                try:
                                    app_date = datetime.strptime(date_str, fmt).date()
                                    break
                                except (ValueError, TypeError):
                                    continue
                            if app_date is None:
                                errors.append(f'Row {idx}: 보관 신청일 형식이 올바르지 않습니다: {date_str}')
                                continue
                        elif hasattr(app_date_cell.value, 'date'):
                            app_date = app_date_cell.value.date()
                        elif hasattr(app_date_cell.value, 'year'):
                            # datetime 객체인 경우
                            app_date = app_date_cell.value
                            if not isinstance(app_date, type(datetime.now().date())):
                                app_date = app_date.date()
                        else:
                            errors.append(f'Row {idx}: 보관 신청일을 파싱할 수 없습니다.')
                            continue
                        
                        # 보관 시작일이 있으면 사용하고, 없으면 None으로 설정
                        if '보관 시작일' in header_map:
                            start_date_cell = row[header_map['보관 시작일'] - 1]
                            if start_date_cell.value:
                                if isinstance(start_date_cell.value, str):
                                    date_str = start_date_cell.value.strip()
                                    if date_str:  # 빈 문자열이 아닌 경우만 파싱
                                        date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']
                                        start_date = None
                                        for fmt in date_formats:
                                            try:
                                                start_date = datetime.strptime(date_str, fmt).date()
                                                break
                                            except (ValueError, TypeError):
                                                continue
                                        if start_date is None:
                                            start_date = None  # 파싱 실패 시 None
                                    else:
                                        start_date = None
                                elif hasattr(start_date_cell.value, 'date'):
                                    start_date = start_date_cell.value.date()
                                elif hasattr(start_date_cell.value, 'year'):
                                    start_date = start_date_cell.value
                                    if not isinstance(start_date, type(datetime.now().date())):
                                        start_date = start_date.date()
                                else:
                                    start_date = None
                            else:
                                start_date = None
                        else:
                            start_date = None
                    except Exception as e:
                        errors.append(f'Row {idx}: 날짜 파싱 오류 - {str(e)}')
                        continue
                    
                    special_notes_cell = row[header_map['특이사항'] - 1]
                    special_notes = str(special_notes_cell.value) if special_notes_cell.value else ''
                    
                    asset = Asset(
                        asset_number=asset_number,
                        storage_application_id=int(storage_application_id),
                        application_date=app_date,
                        storage_start_date=start_date,
                        asset_category=category,
                        special_notes=special_notes,
                        status=AssetStatus.STORED
                    )
                    
                    db.session.add(asset)
                    assets_created.append(asset_number)
                    
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
                    continue
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'created_count': len(assets_created),
                'assets': assets_created,
                'errors': errors
            }), 201
            
        except Exception as e:
            db.session.rollback()
            return handle_error(e, "엑셀 파일 파싱 중 오류가 발생했습니다.")
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@assets_bp.route('/list', methods=['GET'])
def list_assets():
    """보관 자산 목록 조회"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        storage_application_id = request.args.get('storage_application_id')
        
        query = Asset.query.join(StorageApplication).filter(
            StorageApplication.user_id == user.get('id')
        )
        
        if storage_application_id:
            query = query.filter(Asset.storage_application_id == storage_application_id)
        
        assets = query.all()
        
        return jsonify({
            'success': True,
            'assets': [asset.to_dict() for asset in assets]
        }), 200
        
    except Exception as e:
        return handle_error(e, "자산 목록 조회 중 오류가 발생했습니다.")

@assets_bp.route('/<int:asset_id>', methods=['GET'])
def get_asset(asset_id):
    """보관 자산 상세 조회"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        asset = Asset.query.get_or_404(asset_id)
        application = StorageApplication.query.get(asset.storage_application_id)
        
        if application.user_id != user.get('id'):
            return jsonify({'error': 'Unauthorized'}), 403
        
        return jsonify({
            'success': True,
            'asset': asset.to_dict()
        }), 200
        
    except Exception as e:
        return handle_error(e, "자산 상세 조회 중 오류가 발생했습니다.")

@assets_bp.route('/<int:asset_id>', methods=['PUT'])
def update_asset(asset_id):
    """보관 자산 수정"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        asset = Asset.query.get_or_404(asset_id)
        application = StorageApplication.query.get(asset.storage_application_id)
        
        if application.user_id != user.get('id'):
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        if 'application_date' in data:
            asset.application_date = datetime.strptime(data['application_date'], '%Y-%m-%d').date()
        if 'storage_start_date' in data:
            asset.storage_start_date = datetime.strptime(data['storage_start_date'], '%Y-%m-%d').date()
        if 'asset_category' in data:
            try:
                asset.asset_category = AssetCategory(data['asset_category'])
            except ValueError:
                return jsonify({'error': 'Invalid asset category'}), 400
        if 'special_notes' in data:
            asset.special_notes = data['special_notes']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'asset': asset.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return handle_error(e, "자산 수정 중 오류가 발생했습니다.")

@assets_bp.route('/<int:asset_id>', methods=['DELETE'])
def delete_asset(asset_id):
    """보관 자산 삭제"""
    try:
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required'}), 401
        
        asset = Asset.query.get_or_404(asset_id)
        application = StorageApplication.query.get(asset.storage_application_id)
        
        if application.user_id != user.get('id'):
            return jsonify({'error': 'Unauthorized'}), 403
        
        db.session.delete(asset)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Asset deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

